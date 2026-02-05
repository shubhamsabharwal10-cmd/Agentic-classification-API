"""
Excel to JSON DSS Rules Converter

This tool converts DSS rules from Excel format to the JSON format
used by the classification pipeline.

Usage:
    python excel_to_json_converter.py --excel path/to/rules.xlsx --output app/config/dss_rules.json
"""

import openpyxl
import json
import re
import argparse
from typing import Dict, List, Optional, Any
from pathlib import Path


# ============================================================================
# SECTOR MAPPING
# ============================================================================
SECTOR_MAPPING = {
    "MIN1": "mining",
    "IND1": "industry", 
    "INFRA1": "infrastructure",
    # Add more mappings as needed
}

# ============================================================================
# THRESHOLD ATTRIBUTE MAPPING
# ============================================================================
ATTRIBUTE_FIELD_MAPPING = {
    # Industry sector
    "production capacity": {
        "cement": "effective_capacity",
        "steel plant": "effective_capacity",
        "thermal power plant": "power_generation_mw",
        "sugar industry": "sugar_crushing_tcd",
    },
    
    # Mining sector
    "mining lease area": "max_mining_area_ha",
    "coal production": "coal_production_mtpA",
    "sand extraction": "sand_extraction_m3_per_year",
    "mineral area": "minor_mineral_area_ha",
    
    # Infrastructure sector
    "road length": "road_length_km",
    "built up area": "built_up_area_sqm",
    "hydro capacity": "hydro_capacity_mw",
    "dam height": "dam_height_m",
    "power generation": "power_generation_mw",
}


# ============================================================================
# ACTIVITY NAME NORMALIZATION
# ============================================================================
ACTIVITY_NORMALIZATION = {
    # Industry
    "primary metallurgical industry": "steel plant",
    "sponge iron manufacturing": "steel plant",
    "cement plant": "cement",
    "cement manufacturing": "cement",
    "thermal power": "thermal power plant",
    "sugar mill": "sugar industry",
    "sugar manufacturing": "sugar industry",
    
    # Mining
    "iron ore mining": "iron ore",
    "coal mining": "coal mining",
    "sand mining": "sand mining",
    "stone quarry": "stone quarry",
    
    # Infrastructure
    "highway": "highway project",
    "road project": "highway project",
    "construction": "construction project",
    "building project": "construction project",
    "hydroelectric": "hydroelectric project",
    "hydro project": "hydroelectric project",
    "airport": "airport project",
    "port": "port project",
}


# ============================================================================
# CONDITION PARSER
# ============================================================================
class ConditionParser:
    """Parse category condition strings from Excel into JSON rule format"""
    
    @staticmethod
    def parse_condition(condition_str: str, field: str, unit: str) -> Optional[Dict]:
        """
        Parse condition strings like:
        - ">=0.06" → {"field": "...", "op": ">=", "value": 0.06}
        - ">=0.03 and <0.06" → Complex range condition
        - ">0" → {"field": "...", "op": ">", "value": 0}
        - "-" or "NA" or None → No condition (fallback)
        """
        if not condition_str or condition_str in ['-', '--', 'NA', 'None']:
            return None
            
        condition_str = str(condition_str).strip()
        
        # Handle range conditions (e.g., ">=0.03 and <0.06")
        if ' and ' in condition_str.lower():
            parts = re.split(r'\s+and\s+', condition_str, flags=re.IGNORECASE)
            if len(parts) == 2:
                # This is a range - we'll use the lower bound as primary condition
                lower = ConditionParser._parse_simple_condition(parts[0], field)
                return lower
        
        # Handle simple conditions
        return ConditionParser._parse_simple_condition(condition_str, field)
    
    @staticmethod
    def _parse_simple_condition(condition_str: str, field: str) -> Optional[Dict]:
        """Parse simple condition like '>=0.06' or '>5'"""
        # Match operator and value
        match = re.match(r'^\s*([><=]+)\s*([0-9.]+)\s*$', condition_str)
        if match:
            operator = match.group(1)
            value = float(match.group(2))
            return {
                "field": field,
                "op": operator,
                "value": value
            }
        return None


# ============================================================================
# EXCEL TO JSON CONVERTER
# ============================================================================
class ExcelToJSONConverter:
    """Convert Excel DSS rules to JSON format"""
    
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.wb = openpyxl.load_workbook(excel_path)
        self.ws = self.wb.active
        self.rules = {}
        
    def convert(self) -> Dict:
        """Main conversion logic"""
        # Read headers
        headers = [cell.value for cell in self.ws[1]]
        
        # Process each row
        for row_idx in range(2, self.ws.max_row + 1):
            row_data = {}
            for col_idx, cell in enumerate(self.ws[row_idx]):
                if col_idx < len(headers) and headers[col_idx]:
                    row_data[headers[col_idx]] = cell.value
            
            # Skip empty rows
            if not row_data.get('Activity') or not row_data.get('Sub Activity'):
                continue
                
            # Skip NA rules
            if row_data.get('cat A') == 'NA' and row_data.get('cat B1') == 'NA':
                continue
            
            # Process this rule
            self._process_rule(row_data)
        
        return self.rules
    
    def _process_rule(self, row_data: Dict):
        """Process a single rule from Excel row"""
        # Get sector
        sector_code = row_data.get('Sector', '').strip()
        sector = SECTOR_MAPPING.get(sector_code, sector_code.lower() if sector_code else 'unknown')
        
        if sector not in self.rules:
            self.rules[sector] = {}
        
        # Use Sub Activity directly (clean it but don't over-normalize)
        activity_raw = row_data.get('Sub Activity', '').strip()
        activity = self._normalize_activity(activity_raw)
        
        if not activity:
            return
        
        if activity not in self.rules[sector]:
            self.rules[sector][activity] = []
        
        # Get field name
        threshold_attr = row_data.get('Threshold Attribute', '').strip().lower()
        field = self._get_field_name(threshold_attr, activity)
        
        if not field:
            # Use default field for the sector
            if sector == "industry":
                field = "effective_capacity"
            else:
                return
        
        # Parse conditions for each category
        unit = row_data.get('Units', '').strip()
        
        # Determine which categories have conditions
        cat_a_condition = row_data.get('cat A')
        cat_b1_condition = row_data.get('cat B1')
        cat_b2_condition = row_data.get('cat B2')
        
        has_cat_a = cat_a_condition and str(cat_a_condition).strip() not in ['-', '--', 'NA', 'None', '']
        has_cat_b1 = cat_b1_condition and str(cat_b1_condition).strip() not in ['-', '--', 'NA', 'None', '']
        has_cat_b2 = cat_b2_condition and str(cat_b2_condition).strip() not in ['-', '--', 'NA', 'None', '']
        
        # Category A
        if has_cat_a:
            condition = ConditionParser.parse_condition(str(cat_a_condition), field, unit)
            if condition:  # Only add if valid condition parsed
                reason = f"{activity} - Category A"
                self.rules[sector][activity].append({
                    "condition": condition,
                    "category": "A",
                    "reason": reason
                })
        
        # Category B1
        if has_cat_b1:
            condition = ConditionParser.parse_condition(str(cat_b1_condition), field, unit)
            if condition:  # Only add if valid condition parsed
                reason = f"{activity} - Category B1"
                self.rules[sector][activity].append({
                    "condition": condition,
                    "category": "B1",
                    "reason": reason
                })
        
        # Category B2 (usually fallback without condition)
        if has_cat_b2:
            reason = f"{activity} - Category B2"
            # B2 is usually fallback without specific condition
            self.rules[sector][activity].append({
                "category": "B2",
                "reason": reason
            })
        
        # If no valid rules were added for this activity, remove it
        if not self.rules[sector][activity]:
            del self.rules[sector][activity]
    
    def _normalize_activity(self, activity_raw: str) -> Optional[str]:
        """Normalize activity name to match existing pipeline conventions"""
        if not activity_raw:
            return None
            
        activity_lower = activity_raw.lower().strip()
        
        # Check exact matches first
        for key, normalized in ACTIVITY_NORMALIZATION.items():
            if key in activity_lower:
                return normalized
        
        # Return original cleaned version (preserve casing for readability)
        # This allows Excel to have its own activity names
        return activity_raw.strip()
    
    def _get_field_name(self, threshold_attr: str, activity: str) -> Optional[str]:
        """Map threshold attribute to field name"""
        # Check if there's a direct mapping
        if threshold_attr in ATTRIBUTE_FIELD_MAPPING:
            mapping = ATTRIBUTE_FIELD_MAPPING[threshold_attr]
            if isinstance(mapping, dict):
                return mapping.get(activity)
            return mapping
        
        # Fallback: derive from attribute name
        if 'capacity' in threshold_attr:
            return 'effective_capacity'
        elif 'area' in threshold_attr:
            return 'max_mining_area_ha'
        elif 'production' in threshold_attr:
            return 'coal_production_mtpA'
        
        return None
    
    def merge_with_existing(self, existing_json_path: str):
        """
        Merge current rules with existing rules from JSON file.
        New rules are added, existing rules are preserved unless they have the same sector+activity.
        """
        if not Path(existing_json_path).exists():
            print(f"⚠️  No existing rules file found at {existing_json_path}, will create new file")
            return
        
        try:
            with open(existing_json_path, 'r') as f:
                existing_rules = json.load(f)
            
            # Track what we're adding/updating
            added_sectors = []
            added_activities = []
            updated_activities = []
            
            # Merge logic: for each sector in new rules
            for sector, activities in self.rules.items():
                if sector not in existing_rules:
                    # New sector - add it completely
                    existing_rules[sector] = activities
                    added_sectors.append(sector)
                else:
                    # Sector exists - merge activities
                    for activity, rules in activities.items():
                        if activity not in existing_rules[sector]:
                            # New activity - add it
                            existing_rules[sector][activity] = rules
                            added_activities.append(f"{sector}/{activity}")
                        else:
                            # Activity exists - update it
                            existing_rules[sector][activity] = rules
                            updated_activities.append(f"{sector}/{activity}")
            
            # Replace self.rules with merged version
            self.rules = existing_rules
            
            # Print merge summary
            print(f"\n📝 Merge Summary:")
            if added_sectors:
                print(f"   ✨ New sectors added: {', '.join(added_sectors)}")
            if added_activities:
                print(f"   ✨ New activities added: {len(added_activities)}")
                for activity in added_activities[:5]:  # Show first 5
                    print(f"      - {activity}")
                if len(added_activities) > 5:
                    print(f"      ... and {len(added_activities) - 5} more")
            if updated_activities:
                print(f"   🔄 Activities updated: {len(updated_activities)}")
                for activity in updated_activities[:5]:  # Show first 5
                    print(f"      - {activity}")
                if len(updated_activities) > 5:
                    print(f"      ... and {len(updated_activities) - 5} more")
            
        except json.JSONDecodeError as e:
            print(f"⚠️  Error reading existing JSON: {e}")
            print(f"   Will overwrite with new rules")
        except Exception as e:
            print(f"⚠️  Unexpected error during merge: {e}")
            print(f"   Will overwrite with new rules")
    
    def save_json(self, output_path: str, merge: bool = False):
        """Save rules to JSON file"""
        if merge:
            self.merge_with_existing(output_path)
        
        with open(output_path, 'w') as f:
            json.dump(self.rules, f, indent=2)
        print(f"✅ Rules saved to {output_path}")
        print(f"📊 Total sectors: {len(self.rules)}")
        for sector, activities in self.rules.items():
            print(f"   - {sector}: {len(activities)} activities")


# ============================================================================
# MAIN EXECUTION
# ============================================================================
def main():
    parser = argparse.ArgumentParser(description='Convert Excel DSS rules to JSON')
    parser.add_argument('--excel', required=True, help='Path to Excel file')
    parser.add_argument('--output', required=True, help='Output JSON file path')
    parser.add_argument('--merge', action='store_true', 
                        help='Merge with existing rules instead of replacing them')
    
    args = parser.parse_args()
    
    # Validate input file exists
    if not Path(args.excel).exists():
        print(f"❌ Error: Excel file not found: {args.excel}")
        return
    
    # Convert
    mode_str = "merging" if args.merge else "converting"
    print(f"🔄 {mode_str.capitalize()} {args.excel} to JSON...")
    converter = ExcelToJSONConverter(args.excel)
    rules = converter.convert()
    
    # Save
    converter.save_json(args.output, merge=args.merge)
    
    # Preview
    print("\n📋 Preview of generated rules:")
    print(json.dumps(converter.rules, indent=2)[:500] + "...")


if __name__ == "__main__":
    main()
